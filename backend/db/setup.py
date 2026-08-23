"""
Database Setup: Converts the XLSX assessment data into a SQLite database.
Run once with: python -m db.setup
"""

import sqlite3
import openpyxl
import os
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

DATA_DIR = os.getenv("DATA_DIR", "../data/AI Agent Assessment - Candidate Pack")
DB_PATH = os.getenv("DB_PATH", "./db/parcelPilot.db")

XLSX_PATH = os.path.join(DATA_DIR, "ParcelPilot_Assessment_Data.xlsx")

# Dataset snapshot time (from README sheet)
DATASET_SNAPSHOT = "2026-08-16 11:00 Asia/Kolkata"


def setup_database():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Load workbook
    wb = openpyxl.load_workbook(XLSX_PATH)

    # === Accounts ===
    cursor.execute("DROP TABLE IF EXISTS accounts")
    cursor.execute("""
        CREATE TABLE accounts (
            account_id TEXT PRIMARY KEY,
            account_name TEXT,
            plan TEXT,
            status TEXT,
            csm TEXT,
            contract_file TEXT,
            premium_support INTEGER,
            notes TEXT
        )
    """)

    ws_accounts = wb["accounts"]
    headers = [cell.value for cell in next(ws_accounts.iter_rows(min_row=1, max_row=1))]
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cursor.execute(
                "INSERT INTO accounts VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (str(row[0] or ''), str(row[1] or ''), str(row[2] or ''), str(row[3] or ''),
                 str(row[4] or ''), str(row[5] or '') if row[5] else None,
                 1 if row[6] else 0, str(row[7] or '') if row[7] else None)
            )

    # === Orders ===
    cursor.execute("DROP TABLE IF EXISTS orders")
    cursor.execute("""
        CREATE TABLE orders (
            order_id TEXT PRIMARY KEY,
            account_id TEXT,
            carrier TEXT,
            status TEXT,
            booked_at TEXT,
            pickup_window_start TEXT,
            pickup_window_end TEXT,
            pickup_actual_at TEXT,
            shipment_fee_inr REAL,
            carrier_fault INTEGER,
            customer_fault INTEGER,
            cancellation_requested_at TEXT,
            notes TEXT
        )
    """)

    ws_orders = wb["orders"]
    for row in ws_orders.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cursor.execute(
                "INSERT INTO orders VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (str(row[0] or ''), str(row[1] or ''), str(row[2] or ''), str(row[3] or ''),
                 str(row[4]) if row[4] else None, str(row[5]) if row[5] else None,
                 str(row[6]) if row[6] else None, str(row[7]) if row[7] else None,
                 float(row[8]) if row[8] else None,
                 1 if row[9] else 0, 1 if row[10] else 0,
                 str(row[11]) if row[11] else None,
                 str(row[12]) if row[12] else None)
            )

    # === Tickets ===
    cursor.execute("DROP TABLE IF EXISTS tickets")
    cursor.execute("""
        CREATE TABLE tickets (
            ticket_id TEXT PRIMARY KEY,
            account_id TEXT,
            created_at TEXT,
            status TEXT,
            subject TEXT,
            description TEXT,
            channel TEXT,
            assigned_to TEXT,
            last_customer_message_at TEXT,
            historical_resolution TEXT
        )
    """)

    ws_tickets = wb["tickets"]
    for row in ws_tickets.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cursor.execute(
                "INSERT INTO tickets VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (str(row[0] or ''), str(row[1] or ''), str(row[2]) if row[2] else None,
                 str(row[3] or ''), str(row[4] or ''), str(row[5] or ''),
                 str(row[6] or ''), str(row[7] or ''),
                 str(row[8]) if row[8] else None,
                 str(row[9]) if row[9] else None)
            )

    # === Escalations (mocked state-changing action target) ===
    cursor.execute("DROP TABLE IF EXISTS escalations")
    cursor.execute("""
        CREATE TABLE escalations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_id TEXT,
            account_id TEXT,
            reason TEXT,
            priority TEXT,
            created_by TEXT,
            created_at TEXT,
            status TEXT DEFAULT 'open'
        )
    """)

    # === Follow-up tasks ===
    cursor.execute("DROP TABLE IF EXISTS followup_tasks")
    cursor.execute("""
        CREATE TABLE followup_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_id TEXT,
            account_id TEXT,
            task TEXT,
            assigned_to TEXT,
            created_by TEXT,
            created_at TEXT,
            due_at TEXT,
            status TEXT DEFAULT 'pending'
        )
    """)

    conn.commit()
    conn.close()
    print(f"[OK] Database created at {DB_PATH}")
    print(f"   Dataset snapshot: {DATASET_SNAPSHOT}")

    # Verify
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for table in ["accounts", "orders", "tickets"]:
        c.execute(f"SELECT COUNT(*) FROM {table}")
        count = c.fetchone()[0]
        print(f"   {table}: {count} rows")
    conn.close()


if __name__ == "__main__":
    setup_database()
